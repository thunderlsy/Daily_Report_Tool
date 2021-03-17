from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import datetime


# ======================================================
# @Author : Chris
# @Time : 4/11/2020 4:00 PM
# @Desc : 处理CYC数据并写入
# ======================================================

class GenerateTestLog(object):
    date_str = datetime.datetime.now().strftime("%Y/%m/%d")
    datetime_obj = datetime.datetime.strptime(date_str, '%Y/%m/%d')
    # today = datetime.date.today()
    # one_day = datetime.timedelta(days=1)
    # date = today - one_day
    # datetime_obj = str(date)
    # 字体类：颜色：蓝 大小：12 字体：Times New Roman
    cell_blue = Font(color="FF0000FF", sz=12.0, name="Times New Roman")
    # 对齐方式：居中
    cell_alignment_center = Alignment(horizontal='center', vertical='center', textRotation=0,
                                      wrapText=None, shrinkToFit=None, indent=0.0, relativeIndent=0.0,
                                      justifyLastLine=None, readingOrder=0.0)
    # 对齐方式：左对齐
    cell_alignment_left = Alignment(horizontal='left', vertical='center', textRotation=0, wrapText=None,
                                    shrinkToFit=None, indent=0.0, relativeIndent=0.0,
                                    justifyLastLine=None, readingOrder=0.0)

    # 边框类
    cell_border = Border(left=Side(border_style='thin', color='FF000000'),
                         right=Side(border_style='thin', color='FF000000'),
                         top=Side(border_style='thin', color='FF000000'),
                         bottom=Side(border_style='thin', color='FF000000')
                         )

    def __init__(self, address):
        self.address = address
        self.workbook = load_workbook(self.address)
        self.test_log_sheet = self.workbook["TestLog"]
        # self.run()

    # 返回sheet名颜色为蓝色的sheet名列表
    def get_all_blue_sheet(self):
        all_sheets_list = self.workbook.sheetnames
        blue_sheet_list = []
        for i in all_sheets_list:
            sheet_tabColor = self.workbook[i].sheet_properties.tabColor
            if not sheet_tabColor:
                continue
            if sheet_tabColor.rgb == "FF0000FF":
                blue_sheet_list.append(i)
        return blue_sheet_list

    # 传入sheet、表中所有fail的行数，写入数据
    def handle_sheet_from_fail_row(self, table_tem, fail_row_list):

        table_row = table_tem.max_row
        table_col = table_tem.max_column

        actual_row = 0
        if len(fail_row_list) > 0:
            for i in range(2, table_row + 1):
                if not table_tem.cell(i, 1).value:
                    actual_row = i
                    break
        else:
            return

        table_type_list = []
        for i in fail_row_list:
            p_value = table_tem.cell(i, 16).value
            cyc_type = [1]
            if "]:1[" in p_value:
                cyc_row_list = p_value[:-2].split(":1")
                len_cyc_list = len(cyc_row_list)
                for ki in range(len_cyc_list):
                    if not ki:
                        continue
                    if cyc_row_list[ki] == cyc_row_list[ki - 1]:
                        cyc_type[-1] += 1
                    else:
                        cyc_type.append(1)
            table_type_list.append(cyc_type)

        print("table_type_list:", table_type_list)
        log_col = self.test_log_sheet.max_row + 1
        while True:
            if self.test_log_sheet.cell(log_col-1, 7).value or self.test_log_sheet.cell(log_col-1, 3).value:
                break
            else:
                log_col -= 1

        # 合并C列单元格
        tem_num = log_col
        to_log_row = 0
        # table_type_list: [[2, 2, 1], [1], [1, 1]]
        # for 2 in [2, 2, 1]
        for i in table_type_list:
            len_tem = len(i)
            to_log_row += len_tem
            # if len_tem > 1:
            #     self.test_log_sheet.merge_cells("C" + str(tem_num) + ":" + "C" + str(tem_num + len_tem - 1))
            # tem_num = tem_num + len_tem
        # log_col_end = log_col + len(fail_row_list) - 1
        log_col_end = log_col + to_log_row - 1

        # 合并A、B列单元格
        # merge_cell_list = ["A", "B", "D"]
        # for i in merge_cell_list:
        #     self.test_log_sheet.merge_cells(i + str(log_col) + ":" + i + str(log_col_end))
        self.test_log_sheet.merge_cells("A" + str(log_col) + ":" + "A" + str(log_col_end))
        self.test_log_sheet.merge_cells("B" + str(log_col) + ":" + "B" + str(log_col_end))

        fail_table_len = len(fail_row_list)
        for i in range(fail_table_len):
            if not i:
                # A列
                self.test_log_sheet.cell(log_col, 1).value = self.datetime_obj
                self.test_log_sheet.cell(log_col, 1).font = self.cell_blue
                self.test_log_sheet.cell(log_col, 1).number_format = 'm/d;@'
                self.test_log_sheet.cell(log_col, 1).alignment = self.cell_alignment_center
                self.test_log_sheet.cell(log_col, 1).border = self.cell_border

                # B列
                self.test_log_sheet.cell(log_col, 2).value = table_tem.title
                self.test_log_sheet.cell(log_col, 2).font = self.cell_blue
                self.test_log_sheet.cell(log_col, 2).alignment = self.cell_alignment_center
                self.test_log_sheet.cell(log_col, 2).border = self.cell_border

                # D列
                # self.test_log_sheet.cell(log_col, 4).value = actual_row - 1
                # self.test_log_sheet.cell(log_col, 4).font = self.cell_blue
                # self.test_log_sheet.cell(log_col, 4).alignment = self.cell_alignment_center
                # self.test_log_sheet.cell(log_col, 4).border = self.cell_border

            O_split_list = []
            P_split_list = []
            N_split_list = []
            Q_split_list = []
            if len(table_type_list[i]) > 1 or (len(table_type_list[i]) == 1 and table_type_list[i][0] > 1):
                O_value = table_tem.cell(fail_row_list[i], 15).value[:-2]
                O_split_list = O_value.split(":1")
                P_value = table_tem.cell(fail_row_list[i], 16).value[:-2]
                P_split_list = P_value.split(":1")
                N_value = table_tem.cell(fail_row_list[i], 14).value[:-2]
                N_split_list = N_value.split(":1")
                Q_value = table_tem.cell(fail_row_list[i], 17).value[:-2]
                Q_split_list = Q_value.split(":1")
            # elif len(table_type_list[i]) == 1 and table_type_list[i][0] > 1:
            #     O_value = table_tem.cell(fail_row_list[i], 15).value[:-2]
            #     O_split_list = O_value.split(":1")
            #     P_value = table_tem.cell(fail_row_list[i], 16).value[:-2]
            #     P_split_list = P_value.split(":1")
            #     N_value = table_tem.cell(fail_row_list[i], 14).value[:-2]
            #     N_split_list = N_value.split(":1")
            #     Q_value = table_tem.cell(fail_row_list[i], 17).value[:-2]
            #     Q_split_list = Q_value.split(":1")
            # else:
            #     O_split_list[0] = table_tem.cell(fail_row_list[i], 15).value
            #     N_split_list[0] = table_tem.cell(fail_row_list[i], 14).value
            #     Q_split_list[0] = table_tem.cell(fail_row_list[i], 17).value

            # table_type_list: [[2, 2, 1], [1], [1, 1]]
            # for 2 in [2, 2, 1]
            index_begin = 0
            for fail_cyc_row in table_type_list[i]:

                # ABCDEFGHI列边框、字体颜色设置
                for j in range(1, 10):
                    self.test_log_sheet.cell(log_col, j).border = self.cell_border
                    self.test_log_sheet.cell(log_col, j).font = self.cell_blue

                # C列
                if len(table_type_list[i]) > 1:
                    P_to_log = P_split_list[fail_cyc_row + index_begin - 1]
                else:
                    # 同个CYC多个Fail
                    if fail_cyc_row > 1:
                        P_to_log = P_split_list[0]
                    else:
                        P_to_log = table_tem.cell(fail_row_list[i], 16).value
                P_to_log = P_to_log[1:-1]
                print("P_to_log:", P_to_log)
                self.test_log_sheet.cell(log_col, 3).value = P_to_log
                self.test_log_sheet.cell(log_col, 3).alignment = self.cell_alignment_left

                # D列 + F列
                self.test_log_sheet.cell(log_col, 4).value = actual_row - 1
                F_with_1 = "1F/"
                for j in range(table_col, 18, -1):
                    if table_tem.cell(1, j).value == P_to_log:
                        D_row = 0
                        for k in range(2, table_row+1):
                            D_tem_value = table_tem.cell(k, j).value
                            if D_tem_value == "COF" or D_tem_value == "Pass" or D_tem_value == "Fail":
                                D_row += 1
                        self.test_log_sheet.cell(log_col, 4).value = D_row
                        F_value = F_with_1 + str(D_row)
                        self.test_log_sheet.cell(log_col, 6).value = F_value
                        break
                self.test_log_sheet.cell(log_col, 4).alignment = self.cell_alignment_center
                self.test_log_sheet.cell(log_col, 6).alignment = self.cell_alignment_center

                # G列
                self.test_log_sheet.cell(log_col, 7).value = table_tem.cell(fail_row_list[i], 2).value
                self.test_log_sheet.cell(log_col, 7).alignment = self.cell_alignment_left

                # H列
                self.test_log_sheet.cell(log_col, 8).value = table_tem.cell(fail_row_list[i], 1).value
                self.test_log_sheet.cell(log_col, 8).alignment = self.cell_alignment_left

                # I列
                if len(table_type_list[i]) > 1:
                    N_to_log = ((N_split_list[fail_cyc_row + index_begin - 1] + ":1") * fail_cyc_row)[:-2]
                    Q_to_log = ((Q_split_list[fail_cyc_row + index_begin - 1] + ":1") * fail_cyc_row)[:-2]
                    O_to_log = O_split_list[fail_cyc_row + index_begin - 1]
                else:
                    N_to_log = table_tem.cell(fail_row_list[i], 14).value
                    Q_to_log = table_tem.cell(fail_row_list[i], 17).value
                    O_to_log = table_tem.cell(fail_row_list[i], 15).value
                print("O_to_log:", O_to_log, "N_to_log:", N_to_log, "Q_to_log:", Q_to_log)

                # 同个cyc中多个Fail的情况
                # if ":1" in O_to_log:
                if fail_cyc_row > 1:
                    # I_value = ""
                    O_split_I = O_to_log[:-2].split(":1")
                    N_split_I = N_to_log[:-2].split(":1")
                    Q_split_I = Q_to_log[:-2].split(":1")
                    I_value = O_split_I[0]
                    for split_num in range(len(O_split_I)):
                        # I_split = O_split_I[split_num] + N_split_I[split_num] + Q_split_I[split_num] + ":1"
                        I_split = N_split_I[split_num] + Q_split_I[split_num] + ":1"
                        I_value += I_split
                    self.test_log_sheet.cell(log_col, 9).value = "@" + I_value[:-2]
                else:
                    self.test_log_sheet.cell(log_col, 9).value = "@" + O_to_log + N_to_log + Q_to_log
                self.test_log_sheet.cell(log_col, 9).alignment = self.cell_alignment_left

                index_begin += fail_cyc_row
                log_col += 1

        # C、D、F列合并相同项
        # [[1, 2], [5, 8], [11, 14]]
        C_merge_list = []
        for i in range(tem_num, log_col_end+1):

            if len(C_merge_list) > 0 and C_merge_list[-1][0] <= i <= C_merge_list[-1][1]:
                continue

            k = 1
            while True:
                if self.test_log_sheet.cell(i, 3).value == self.test_log_sheet.cell(i+k, 3).value:
                    C_merge_list.append([i, i+k])
                    k += 1
                else:
                    if k > 1:
                        for j in range(k-2):
                            del C_merge_list[-2]
                    break

        print("C_merge_list:", C_merge_list)
        print("tem_num:", tem_num, "log_col_end:", log_col_end)
        for i in C_merge_list:
            F_str_to_list = list(self.test_log_sheet.cell(i[0], 6).value)
            F_str_to_list[0] = str(i[1] - i[0] + 1)
            self.test_log_sheet.cell(i[0], 6).value = ''.join(F_str_to_list)
            self.test_log_sheet.merge_cells(start_row=i[0], start_column=3, end_row=i[1], end_column=3)
            self.test_log_sheet.merge_cells(start_row=i[0], start_column=4, end_row=i[1], end_column=4)
            self.test_log_sheet.merge_cells(start_row=i[0], start_column=6, end_row=i[1], end_column=6)

    def do_C_same_value(self, i):
        if self.test_log_sheet.cell(i, 3).value == self.test_log_sheet.cell(i + 1, 3).value:
            pass

    def save_book(self):
        self.workbook.save(self.address)

    def get_fail_row(self, table_tem):
        fail_row_list = []
        k = 2
        while True:
            cell_d_now = table_tem.cell(k, 4).value
            cell_m_now = table_tem.cell(k, 13).value
            if cell_d_now is None:
                return fail_row_list
            if cell_d_now == self.datetime_obj and (cell_m_now == "Fail" or cell_m_now == "COF"):
                fail_row_list.append(k)
            k += 1
        pass

    # 处理每个sheet表
    def run(self):
        blue_list = self.get_all_blue_sheet()
        print("blue_list:", blue_list)
        for i in blue_list:
            table_tem = self.workbook[i]
            try:
                # 传入sheet，查找D列是否为当天日期，M列是否为Fail与COF，返回符合的row列表
                fail_row_list = self.get_fail_row(table_tem)
                print("table_tem:", table_tem, "fail_row_list:", fail_row_list)
                # 传入sheet、fail行数 写入数据
                self.handle_sheet_from_fail_row(table_tem, fail_row_list)
            except:
                pass

# sheet颜色        蓝色：FF0000FF 红色：FFFC0107
# cell表格填充颜色   蓝色：FF0033CC 紫色：FF7030A0 黄色：FFFFFFCC 红色：FFFFFFCC
# cell表格字体颜色   蓝色：FF0000FF
# kkk = GenerateTestLog("fox.xlsx")
# kkk.run()
# kkk.save_workbook_1("Cat.xlsx")
