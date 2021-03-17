import time
import openpyxl
import datetime
from generate_test_log import GenerateTestLog


# ======================================================
# @Author : Esther
# @Time : 28/10/2020 5:00 PM
# @Desc : 打开细项文件，将其更新在dailyreport文件中，更新过的sheet标签颜色改为蓝色，使用更新当天日期命名dailyreport文件
# ======================================================
class OpenMulFile:

    def __init__(self, filepathlist, dailyreportpath):
        self.filepathlist = filepathlist
        self.daily_report_path = dailyreportpath
        self.filepathdic = {}
        self.daily_report = None

    # 获取同名sheet列表
    def get_same_sheet_name(self):
        # today = datetime.date.today()
        # oneday = datetime.timedelta(days=1)
        # date = today - oneday
        # date_str = str(date)
        date = datetime.datetime.now().strftime('%m%d%Y')
        date_str = str(date[-4:]) + '-' + str(date[:-6]) + '-' + str(date[-6:-4])
        same_names_dic = {}
        for key, value in self.filepathdic.items():
            same_sheet_names = []
            for file_sheet_name in value.sheetnames:
                if file_sheet_name in self.daily_report.sheetnames:
                    max_row = value[file_sheet_name].max_row
                    for i in range(max_row):
                        i += 1
                        # 如果有更新當天的日期，就append這個sheet名到列表中，然後執行後續操作
                        if date_str in str(value[file_sheet_name]['D' + str(i)].value):
                            same_sheet_names.append(file_sheet_name)
                            same_names_dic[value] = same_sheet_names
                            break

        return same_names_dic

    # 删除daily_report中某个sheet的内容
    def del_daily_report_same_sheet(self, sheet_name):
        daily_sheet = self.daily_report[sheet_name]
        daily_max_row = daily_sheet.max_row
        daily_max_col = daily_sheet.max_column
        for row in range(1, daily_max_row + 1):
            for col in range(1, daily_max_col + 1):
                daily_sheet.cell(row, col).value = None
        # return self.daily_report

    # 若daily_report中存在与fail同名的sheet,将file中sheet的内容赋给daily_report与他同名sheet
    def give_value(self):

        for filepath in self.filepathlist:
            value = openpyxl.load_workbook(filepath)
            self.filepathdic[filepath] = value

        # 读取dailyreport文件
        self.daily_report = openpyxl.load_workbook(self.daily_report_path)

        same_names_dic = self.get_same_sheet_name()
        for key, value in same_names_dic.items():
            for same_sheet_name in value:
                # 删除daily_report中与细项表中同名sheet的内容
                self.del_daily_report_same_sheet(same_sheet_name)
                # 读取file和daily_report中同名的sheet
                file_sheet = key[same_sheet_name]
                daily_sheet = self.daily_report[same_sheet_name]
                # 获取file_sheet最大行、列数
                file_sheet_max_row = file_sheet.max_row
                file_sheet_max_col = file_sheet.max_column
                # 将fail中sheet的值赋给daily_report与他同名sheet
                for row in range(1, file_sheet_max_row + 1):
                    for col in range(1, file_sheet_max_col + 1):
                        daily_sheet.cell(row, col).value = file_sheet.cell(row, col).value
                # 更新sheet标签颜色
                self.daily_report[same_sheet_name].sheet_properties.tabColor = "FF0000FF"
        save_path = self.save_daily_report()
        return save_path
        # GenerateTestLog(save_path)

    # 保存文件
    def save_daily_report(self):
        # 以日期命名daily report文件
        daily_name = self.daily_report_path[:-5] + datetime.datetime.now().strftime('%m%d%Y') + '.xlsx'
        # daily_name = 'daily_test4.xlsx'
        self.daily_report.save(daily_name)
        return daily_name

# filepaths = ['/Users/x2000467/Desktop/xixiang1.xlsx','/Users/x2000467/Desktop/xixaing2.xlsx']
# dailyreportpath = '/Users/x2000467/Desktop/daily.xlsx'
# excel = OpenMulFile(filepathlist=filepaths,dailyreportpath=dailyreportpath)
# excel.give_value()


